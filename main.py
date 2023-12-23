from openpyxl import Workbook

#Makes sure there is a valid workbook working
def CheckWorkbookExists():
    from os import path
    from openpyxl import load_workbook

    if path.exists("./JobListings.xlsx"):
        currWB = load_workbook(filename = "JobListings.xlsx")
    else:
        #Creates new workbook template
        currWB = Workbook()
        currWB["Sheet"].title = "Indeed_Listings"
        currSheet = currWB["Indeed_Listings"]

        currSheet['A1'] = "COMPANY"
        currSheet['B1'] = "POSITION"
        currSheet['C1'] = "LOCATION"
        currSheet['D1'] = "EXTRA"
        currSheet['E1'] = "LINK"

    return currWB

def ReadConfigFile():
    with open("config.txt", "rt") as fin:
        fin.readline() #garbage line
        inputLink = fin.readline()
        fin.readline()
        fin.readline()
        filterListStr = fin.readline()

        if (filterListStr != "None"):
            filterList = filterListStr.split(",")
            filterList = [item.strip() for item in filterList]
        else:
            filterList = None

    return inputLink[:-1], filterList  #[:-1] removes new line

def matchFilter(keyWords, titleStr):
    import re

    for keyWord in keyWords:
        regexMatch = re.search(f"({keyWord})+", titleStr)
        if (regexMatch):
            break

    return regexMatch

def ImportJobInfo(currPg, inputLink, WB, currRowPos, newWidthDimensions, ExistJobs, keyWords):
    from bs4 import BeautifulSoup
    from selenium import webdriver

    widthDimensions = newWidthDimensions #The current width dimensions are updated from the last width dimensions
    counter = currRowPos #CurrentRowPos is used as the counter for the jobOpening loop

    #Current sheet
    FirstSheet = WB["Indeed_Listings"]


    #Opens browser, pulls its html, and constructs a beautiful soup object in lxml form 
    browser = webdriver.Chrome()
    browser.get(inputLink)
    soup = BeautifulSoup(browser.page_source,"lxml")

    browser.close()

    #Finds each job opening div
    jobOpenings = soup.find_all("div",class_ = "cardOutline")
    
    #Column letters
    itemLetters = ["A", "B", "C", "D", "E"]

    noNewJobs = True
    for jobOpening in jobOpenings: 
        companyName = jobOpening.find("span", class_ = "css-1x7z1ps").text
        jobTitle = jobOpening.find(class_ = "jobTitle").text
        location = jobOpening.find("div", class_ = "css-t4u72d").text
        extraInfo = jobOpening.find(class_="css-1ihavw2")
        link = "www.indeed.com" + jobOpening.find("a")["href"]

        #Used to set each attribute in its correct column (position matters)
        itemAttr = [companyName, jobTitle, location, extraInfo, link]

        if (jobTitle in ExistJobs): #Makes sure that no jobs are duplicated
            pass
        else:
            if (keyWords != None):
                strMatches = matchFilter(keyWords, jobTitle)    #Uses regex to find if the job is in the filter

            if ((keyWords == None) or strMatches):  #If none, no filter is required, else check if its in the filter
                for letter,attribute in zip(itemLetters, itemAttr):
                    if (attribute != None):
                        #Some of the extra info attributes are None (its text must be acquired here)
                        if (letter == "D"):
                            FirstSheet[letter + str(counter)] = attribute.text
                            continue

                        FirstSheet[letter + str(counter)] = attribute
                        widthDimensions[letter] = max(widthDimensions[letter], len(attribute)) #gets max width
                        noNewJobs = False
                counter = counter + 1
        
    if (noNewJobs == False):
        #Updates width using widthDimensions dict
        for letter in itemLetters:
            FirstSheet.column_dimensions[letter].width = widthDimensions[letter]


    #Prepares the current row position, the next page and the next page link for the following iteration
    currRowPos = counter
    nextPage = currPg + 1
    nextLink = "https://www.indeed.com" + soup.find(attrs = {"aria-label" : nextPage})["href"]

    print(f"Page {currPg} scanned")

    return nextLink, currRowPos, widthDimensions


def createListExistingJobs(WB):
    #Current sheet
    FirstSheet = WB["Indeed_Listings"]

    #Prepare a list of all job titles from the excel sheet
    jobList = []
    counter = 2 #Two is the starting pos for actual job entrees
    tempStrInput = "Not Empty"
    while(tempStrInput is not None):
        tempStrInput = FirstSheet["B" + str(counter)].value
        jobList.append(tempStrInput)
        counter = counter + 1

    #Remove None
    jobList.pop()

    return jobList, counter-1 

if (__name__ == "__main__"):

    wb = CheckWorkbookExists()

    #Default/Initial values
    nextLink, keyWords = ReadConfigFile()
    inputWidthDims = {"A" : 0, "B" : 0, "C" : 0, "D" : 0, "E" : 0}

    ExistingJobs, rowPos = createListExistingJobs(wb)   #Creates a list of existing jobs on excel sheet in order to find what not to add and where to append

    for pgNum in range(1,4):
        nextLink, rowPos, inputWidthDims = ImportJobInfo(pgNum, nextLink, wb, rowPos, inputWidthDims, ExistingJobs, keyWords)

    print("Job_Listings.xlsx updated")

    wb.save(filename = "JobListings.xlsx")


