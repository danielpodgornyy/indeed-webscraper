from openpyxl import Workbook

item = Workbook()
firstPg = item['Sheet']
firstPg.title = "lol"
#newItem = item.create_sheet("Jabroney")

print(item.sheetnames)

item.save("lol.xlsx")